#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2022/7/5 9:44
# @Author : ma.fei
# @File : t_sql_export.py
# @Software: PyCharm

import datetime
import os
import re
import traceback
import zipfile

import openpyxl
import xlwt

from web.model.t_dmmx import get_dmmc_from_dm
from web.model.t_ds import get_ds_by_dsid
from web.model.t_sql import exe_query_exp
from web.model.t_sql_release import send_wx,send_qywx
from web.model.t_user import get_user_by_loginame
from web.utils.common import format_sql as fmt_sql
from web.utils.common import send_mail_param, get_sys_settings, current_time, current_rq
from web.utils.mysql_async import async_processer

def set_header_styles(p_fontsize, p_color):
    header_borders = xlwt.Borders()
    header_styles = xlwt.XFStyle()
    # add table header style
    header_borders.left = xlwt.Borders.THIN
    header_borders.right = xlwt.Borders.THIN
    header_borders.top = xlwt.Borders.THIN
    header_borders.bottom = xlwt.Borders.THIN
    header_styles.borders = header_borders
    header_pattern = xlwt.Pattern()
    header_pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    header_pattern.pattern_fore_colour = p_color
    # add font
    font = xlwt.Font()
    font.name = u'微软雅黑'
    font.bold = True
    font.size = p_fontsize
    header_styles.font = font
    # add alignment
    header_alignment = xlwt.Alignment()
    header_alignment.horz = xlwt.Alignment.HORZ_CENTER
    header_alignment.vert = xlwt.Alignment.VERT_CENTER
    header_styles.alignment = header_alignment
    header_styles.borders = header_borders
    header_styles.pattern = header_pattern
    return header_styles


async def save_exp_sql(p_dbid, p_cdb, p_sql, p_flag, p_user):
    result = {}
    try:
        sql = """insert into t_sql_export(dbid,db,sqltext,status,creation_date,creator,last_update_date,updator) 
                 values('{0}','{1}',"{2}",'{3}','{4}','{5}','{6}','{7}')
            """.format(p_dbid, p_cdb, fmt_sql(p_sql), p_flag,
                       current_time(), p_user['login_name'],
                       current_time(), p_user['login_name'])

        await async_processer.exec_sql(sql)
        result['code'] = '0'
        result['message'] = '发布成功！'
        return result
    except:
        traceback.print_exc()
        result['code'] = '-1'
        result['message'] = '发布失败!'
        return result


async def update_exp_sql(p_dbid, p_cdb, p_sql, p_flag, p_user, p_id):
    result = {}
    try:
        st = """update t_sql_export
               set dbid={},
                   db='{}',
                   sqltext='{}',
                   status='{}',
                   updator='{}',
                   last_update_date=now() 
                where id={}
           """.format(p_dbid, p_cdb, fmt_sql(p_sql), p_flag, p_user['login_name'], p_id)
        await async_processer.exec_sql(st)
        result['code'] = '0'
        result['message'] = '更新成功！'
        return result
    except:
        traceback.print_exc()
        result['code'] = '-1'
        result['message'] = '更新失败!'
        return result


async def export_query(p_dbid, p_creator, p_key, p_username, p_userid):
    v_where = ''
    if p_creator != '':
        v_where = v_where + " and a.creator='{}'\n".format(p_creator)

    if p_dbid != '':
        v_where = v_where + " and a.dbid='{}'\n".format(p_dbid)
    else:
        v_where = v_where + """ and exists(select 1 from t_user_proj_privs x 
                                           where x.proj_id=b.id and x.user_id='{0}' and priv_id='6')""".format(p_userid)

    if p_username != 'admin':
        v_where = v_where + "  and  a.creator='{0}'".format(p_username)

    if p_key != '':
        v_where = v_where + " and a.sqltext like '%{}%'\n".format(p_key)

    sql = """SELECT  a.id, 
                     b.db_desc,
                     a.db,
                     (SELECT NAME FROM t_user e WHERE e.login_name=a.creator) creator,
                     DATE_FORMAT(a.creation_date,'%Y-%m-%d %h:%i:%s')  creation_date,
                     (SELECT NAME FROM t_user e WHERE e.login_name=a.auditor) auditor,
                     DATE_FORMAT(a.audit_date,'%y-%m-%d %h:%i:%s')   audit_date,
                     c.dmmc as status      
                FROM t_sql_export a,t_db_source b,t_dmmx c
                WHERE a.dbid=b.id
                AND c.dm='41'
                AND a.status=c.dmm
                {} ORDER BY a.creation_date desc""".format(v_where)
    print(sql)
    return await async_processer.query_list(sql)


async def exp_data(static_path, p_ds, p_sql):
    row_data = 0
    workbook = xlwt.Workbook(encoding='utf8')
    worksheet = workbook.add_sheet('kpi')
    header_styles = set_header_styles(45, 1)
    os.system('cd {0}'.format(static_path + '/downloads/sql'))
    file_name = static_path + '/downloads/sql/exp_sql_{0}.xls'.format(current_rq())
    file_name_s = 'exp_sql_{0}.xls'.format(current_rq())
    sql_header = """select * from ({}) x limit 1""".format(p_sql)

    # 写表头
    desc = await async_processer.query_one_desc_by_ds(p_ds, sql_header)
    for k in range(len(desc)):
        worksheet.write(row_data, k, desc[k][0], header_styles)
        if k in (3, 5):
            worksheet.col(k).width = 8000
        else:
            worksheet.col(k).width = 4000

    # 循环项目写单元格
    row_data = row_data + 1
    rs3 = await async_processer.query_list_by_ds(p_ds, p_sql)
    for i in rs3:
        for j in range(len(i)):
            if i[j] is None:
                worksheet.write(row_data, j, '')
            else:
                worksheet.write(row_data, j, str(i[j]))
        row_data = row_data + 1

    workbook.save(file_name)
    print("{0} export complete!".format(file_name))

    # 生成zip压缩文件
    zip_file = static_path + '/downloads/port/exp_kpi_{0}.zip'.format(current_rq())
    rzip_file = '/static/downloads/port/exp_kpi_{0}.zip'.format(current_rq())

    # 若文件存在则删除
    if os.path.exists(zip_file):
        os.system('rm -f {0}'.format(zip_file))

    z = zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED, allowZip64=True)
    z.write(file_name, arcname=file_name_s)
    z.close()

    # 删除json文件
    os.system('rm -f {0}'.format(file_name))
    return rzip_file


async def query_export(p_id):
    st = """select  a.id,
                    a.dbid, 
                    a.db, 
                    a.sqltext, 
                    a.status
           from t_sql_export a where a.id={}""".format(p_id)
    return await async_processer.query_dict_one(st)


async def delete_export(p_id):
    try:
        st = "delete from t_sql_export where id={}".format(p_id)
        await async_processer.exec_sql(st)
        return {'code': '0', 'message': '删除成功!'}
    except:
        traceback.print_exc()
        return {'code': '-1', 'message': '删除失败!'}


async def get_sql_export(p_id):
    sql = "select * from t_sql_export where id={}".format(p_id)
    return await async_processer.query_dict_one(sql)


async def get_export_params(p_sqlid, p_user, p_status, p_message, p_host):
    """
     功能：工单发布发送邮件及微信参数字典
    """
    wkno = await get_sql_export(p_sqlid)
    p_ds = await get_ds_by_dsid(wkno['dbid'])
    p_ds['service'] = wkno['db']
    email = p_user['email']
    settings = await get_sys_settings()
    v_title = '工单导出审核情况[{}]'.format(wkno['id'])
    nowTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    creator = (await get_user_by_loginame(wkno['creator']))['name']
    cmail = (await get_user_by_loginame(wkno['creator']))['email']
    auditor = (await get_user_by_loginame(wkno['auditor']))['name']
    status = (await get_dmmc_from_dm('41', wkno['status']))[0]

    if p_host == "124.127.103.190":
        p_host = "124.127.103.190:65482"
    elif p_host in ('10.2.39.18', '10.2.39.20', '10.2.39.21'):
        p_host = '{}:81'.format(p_host)
    else:
        p_host = p_host

    send_params = {
        'wkno': wkno,
        'p_ds': p_ds,
        'email': email,
        'settings': settings,
        'title': v_title,
        'nowTime': nowTime,
        'creator': creator,
        'cmail': cmail,
        'auditor':auditor,
        'status': status,
        'reason': '',
        'p_host': p_host,
        'p_user': p_user,
        'msg_type':'EXPORT',
        'to_user' : '{}|{}'.format(p_user['wkno'],
                                   (await get_user_by_loginame(wkno['creator']))['wkno'])
    }
    return send_params

def send_export_mail(params):
    """
         功能：发送工单发布邮件
         参数：工单发布字典信息
    """
    v_html = '''<html>
		<head>
		   <style type="text/css">
			   .xwtable {width: 90%;border-collapse: collapse;border: 1px solid #ccc;}
			   .xwtable thead td {font-size: 12px;color: #333333;
					      text-align: center;background: url(table_top.jpg) repeat-x top center;
				              border: 1px solid #ccc; font-weight:bold;}
			   .xwtable thead th {font-size: 12px;color: #333333;
				              text-align: center;background: url(table_top.jpg) repeat-x top center;
					      border: 1px solid #ccc; font-weight:bold;}
			   .xwtable tbody tr {background: #fff;font-size: 12px;color: #666666;}
			   .xwtable tbody tr.alt-row {background: #f2f7fc;}
			   .xwtable td{line-height:20px;text-align: left;padding:4px 10px 3px 10px;height: 18px;border: 1px solid #ccc;}
		   </style>
		</head>
		<body>
              <table class='xwtable'>
                  <tr><td width="20%">发送时间</td><td width="80%">$$TIME$$</td></tr>
                  <tr><td>数据库名</td><td>$$DBINFO$$</td></tr>
                  <tr><td>提交人员</td><td>$$CREATOR$$</td></tr>
                  <tr><td>审核人员</td><td>$$AUDITOR$$</td></tr>
                  <tr><td>工单类型</td><td>$$TYPE$$</td></tr>
                  <tr><td>工单状态</td><td>$$STATUS$$</td></tr>
                  <tr><td>工单详情</td><td>$$DETAIL$$</td></tr>
              </table>    
		</body>
	    </html>'''
    v_html = v_html.replace('$$TIME$$', params['nowTime'])
    v_html = v_html.replace('$$DBINFO$$',
                              params['p_ds']['url'] + params['p_ds']['service']
                              if params['p_ds']['url'].find(params['p_ds']['service']) < 0 else params['p_ds']['url'])
    v_html = v_html.replace('$$CREATOR$$', params['creator'])
    v_html = v_html.replace('$$AUDITOR$$', params['auditor'])
    v_html = v_html.replace('$$TYPE$$', params['otype'])
    v_html = v_html.replace('$$STATUS$$', params['status'])
    v_html = v_html.replace('$$DETAIL$$', 'http://{}/sql/detail?release_id={}'.format(params['p_host'], params['p_sqlid']))
    v_html = v_html.replace('$$ERROR$$', params.get('error',''))
    send_mail_param(params['settings'].get('send_server'),
                    params['settings'].get('sender'),
                    params['settings'].get('sendpass'),
                    params['email'],
                    params['cmail'],
                    params['v_title'],
                    v_html)

async def upd_exp_sql(p_sqlid, p_user, p_status, p_message, p_host):
    result = {}
    try:
        sql = """update t_sql_export 
                  set  status ='{}' ,
                       last_update_date =now(),
                       updator='{}',
                       audit_date =now() ,
                       auditor='{}',
                       audit_message='{}'
                where id='{}'""".format(p_status, p_user['login_name'], p_user['login_name'], p_message, p_sqlid)
        print(sql)
        await async_processer.exec_sql(sql)

        params = await get_export_params(p_sqlid, p_user, p_status, p_message, p_host)

        # send mail
        # send_export_mail(params)

        # send to wx
        #await send_wx(params)

        # send to qywx
        #await send_qywx(params)

        result['code'] = '0'
        result['message'] = '审核成功!'
        return result
    except:
        traceback.print_exc()
        result['code'] = '-1'
        result['message'] = '审核异常!'
        return result


async def query_exp_audit(p_name, p_dsid, p_creator, p_userid, p_username):
    print('p_creator=', p_creator, 'p_userid', p_username)
    v_where = ''
    if p_name != '':
        v_where = v_where + " and a.sqltext like '%{0}%'\n".format(p_name)

    if p_dsid != '':
        v_where = v_where + " and a.dbid='{0}'\n".format(p_dsid)
    else:
        v_where = v_where + """ and exists(select 1 from t_user_proj_privs x 
                                           where x.proj_id=b.id and x.user_id='{0}' and priv_id='6')""".format(p_userid)
    if p_creator != '':
        v_where = v_where + " and a.creator='{0}'\n".format(p_creator)

    if p_username != 'admin':
        v_where = v_where + " and a.creator='{0}'\n".format(p_username)

    sql = """SELECT  a.id, 
                     b.db_desc,
                     a.db,
                     (SELECT NAME FROM t_user e WHERE e.login_name=a.creator) creator,
                     DATE_FORMAT(a.creation_date,'%Y-%m-%d %h:%i:%s')  creation_date,
                     (SELECT NAME FROM t_user e WHERE e.login_name=a.auditor) auditor,
                     DATE_FORMAT(a.audit_date,'%y-%m-%d %h:%i:%s')   audit_date,
                     c.dmmc as status
            FROM t_sql_export a,t_db_source b,t_dmmx c
            WHERE a.dbid=b.id
              AND c.dm='41'
              AND a.status=c.dmm
              {0}
            order by a.creation_date desc
          """.format(v_where)
    print(sql)
    return await async_processer.query_list(sql)


async def update_export(p_id, p_status, p_process, p_file='', p_real_file='', p_size='', p_error=''):
    st = "update t_sql_export_task a " \
         " set a.status='{}',a.process='{}',a.file='{}',a.real_file='{}',a.size='{}',a.error='{}' where id={}" \
        .format(p_status, p_process, p_file, p_real_file, p_size, p_error, p_id)
    await async_processer.exec_sql(st)


async def export_insert(p_user, p_bbid):
    st = """insert into t_sql_export_task(release_id,status,process,creator,create_date)
                values({},'1','0%','{}',now())""".format(p_bbid, p_user['login_name'])
    id = await async_processer.exec_ins_sql(st)
    return id


async def exp_data_xlsx(static_path, p_bbid, p_data, p_id):
    try:
        row_data = 1
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(index=0, title=p_bbid)
        file_path = static_path + '/downloads/sql'
        os.system('cd {0}'.format(file_path))
        file_name = static_path + '/downloads/sql/exp_bbtj_{}_{}.xlsx'.format(p_bbid, current_rq())
        file_name_s = 'exp_sql_{}_{}.xls'.format(p_bbid, current_rq())
        # write header
        for k in range(len(p_data['column'])):
            ws.cell(column=k + 1, row=row_data, value=p_data['column'][k]['title'])
        await update_export(p_id, '3', '25%')

        # write body
        n_batch_size = 500
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        n_total_rows = len(p_data['data'])
        row_data = row_data + 1
        for i in p_data['data']:
            for j in range(len(i)):
                if i[j] is None:
                    ws.cell(row=row_data, column=j + 1, value='')
                else:
                    v = ILLEGAL_CHARACTERS_RE.sub(r'', str(i[j]))
                    ws.cell(row=row_data, column=j + 1, value=v)
            row_data = row_data + 1
            if row_data % n_batch_size == 0:
                await update_export(p_id, '3', str(round(row_data / 75, 2) * 100) + '%')

        await update_export(p_id, '3', '98%')

        wb.save(file_name)
        print("{0} export complete!".format(file_name))
        zip_file = static_path + '/downloads/sql/exp_sql_{}_{}.zip'.format(p_bbid, current_rq())
        rzip_file = '/static/downloads/sql/exp_sql_{}_{}.zip'.format(p_bbid, current_rq())

        if os.path.exists(zip_file):
            os.system('rm -f {0}'.format(zip_file))

        z = zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED, allowZip64=True)
        z.write(file_name, arcname=file_name_s)
        z.close()

        file_size = os.path.getsize(file_name)
        os.system('rm -f {0}'.format(file_name))

        # update path,file,size
        await update_export(p_id, '3', '100%', rzip_file, zip_file, file_size)
        return rzip_file
    except:
        await update_export(p_id, '4', '0%', '', '', '', traceback.print_exc())


async def query_bbgl_data(p_bbid):
    try:
        cfg = await query_export(p_bbid)
        result = await exe_query_exp(cfg['dbid'], cfg['sqltext'], cfg['db'])
        result = {"data": result['data'], "column": result['column'], "status": result['status'], "msg": result['msg']}
        return result
    except:
        result = {"data": '', "column": '', "status": '1', "msg": traceback.print_exc()}
        return result


async def export_data(p_bbid, p_user, path):
    try:
        id = await export_insert(p_user, p_bbid)
        res = await query_bbgl_data(p_bbid)
        if res['status'] == '1':
            return {"code": -1, "message": res['msg']}
        await update_export(id, '2', '20%')
        zip_file = await exp_data_xlsx(path, p_bbid, res, id)
        return {"code": 0, "message": zip_file}
    except:
        return {"code": -1, "message": traceback.print_exc()}


async def get_download(p_id):
    st = "select *from t_sql_export_task where id={}".format(p_id)
    return await async_processer.query_dict_one(st)


async def del_export(p_id):
    try:
        res = await get_download(p_id)
        print('res=', res)
        os.system('rm -f {0}'.format(res.get('real_file')))
        st = "delete from t_sql_export_task where id={}".format(p_id)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '删除成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '删除失败!'}


async def query_exp_task(p_dsid, p_creater, p_key, p_userid):
    vv = ''
    if p_key != '':
        vv = vv + " and a.sqltext like '%{0}%'\n".format(p_key)

    if p_dsid != '':
        vv = vv + " and a.dbid='{0}'\n".format(p_dsid)
    else:
        vv = vv + """ and exists(select 1 from t_user_proj_privs x 
                                   where x.proj_id=b.dbid and x.user_id='{0}' and priv_id='6')""".format(p_userid)

    if p_creater != '':
        vv = vv + " and a.creator='{0}'\n".format(p_creater)

    st = """SELECT 
                 a.id AS task_id,
                 b.id AS expid,	 
                 SUBSTR(d.dmmc,1,20) AS flag,
                 a.process,
                 a.size,
                 c.name,
                 DATE_FORMAT(a.create_date,'%Y-%m-%d %H:%i:%s')  AS  create_date
            FROM t_sql_export_task a,t_sql_export b,t_user c,t_dmmx d 
            WHERE a.release_id=b.id
              AND a.creator=c.login_name
              AND a.status=d.dmm
              AND d.dm='43'
             {}""".format(vv)
    print('st=', st)
    return await async_processer.query_list(st)
