#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/6/30 15:46
# @Author  : ma.fei
# @File    : t_user.py
# @Software: PyCharm

import datetime
import json
import os
import traceback
import zipfile

import openpyxl
import requests
import xlwt

from web.model.t_ds import get_ds_by_dsid
from web.model.t_sql import exe_query_exp
from web.utils.common import current_rq
from web.utils.common import format_sql
from web.utils.common import format_sql as fmt_sql, get_seconds
from web.utils.mysql_async import async_processer


async def get_config(p_bbdm):
    st = "select * from t_bbgl_config where bbdm='{}'".format(p_bbdm)
    if (await async_processer.query_one(st)) is not None:
        return await async_processer.query_dict_one(st)
    else:
        return {}


async def get_preprocess(p_bbdm):
    st = "select statement,description from t_bbgl_preproccess where bbdm='{}' ORDER BY xh".format(p_bbdm)
    return await async_processer.query_dict_list(st)


async def get_headers(p_bbdm):
    st = "select xh,header_name,header_width from t_bbgl_header where bbdm='{}' ORDER BY xh".format(p_bbdm)
    return await async_processer.query_dict_list(st)


async def get_filter(p_bbdm):
    st = "select  * from t_bbgl_filter where bbdm='{}' ORDER BY xh".format(p_bbdm)
    return await async_processer.query_dict_list(st)


async def get_header_xh(p_bbdm):
    st = "select max(xh)+1 from t_bbgl_header where bbdm='{}'".format(p_bbdm)
    rs = await async_processer.query_one(st)
    print('rs=', rs)
    if rs[0] == None:
        return 1
    else:
        return rs[0]


async def get_filter_xh(p_bbdm):
    st = "select max(xh)+1 from t_bbgl_filter where bbdm='{}'".format(p_bbdm)
    rs = await async_processer.query_one(st)
    print('rs=', rs)
    if rs[0] == None:
        return 1
    else:
        return rs[0]


async def get_preprocess_xh(p_bbdm):
    st = "select max(xh)+1 from t_bbgl_preproccess where bbdm='{}'".format(p_bbdm)
    rs = await async_processer.query_one(st)
    print('rs=', rs)
    if rs[0] == None:
        return 1
    else:
        return rs[0]


async def save_bbgl(bbdm, bbmc, dsid, db, userid):
    try:
        if (await check_bbdm_exists(bbdm)) == 0:
            st = "insert into t_bbgl_config(bbdm,bbmc,dsid,db,creator,create_date,last_update_date) \
                    values('{}','{}','{}','{}','{}',now(),now())".format(bbdm, bbmc, dsid, db, userid)
            await async_processer.exec_sql(st)
            return {'code': 0, 'message': '保存成功!'}
        else:
            st = """update t_bbgl_config set bbmc='{}',dsid='{}',db='{}',last_update_date=now() where bbdm='{}'""".format(
                bbmc, dsid, db, bbdm)
            await async_processer.exec_sql(st)
            return {'code': 0, 'message': '更新成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '保存失败!'}


async def save_bbgl_header(bbdm, name, width):
    try:
        st = "insert into t_bbgl_header(bbdm,xh,header_name,header_width) \
                 values('{}',{},'{}','{}')".format(bbdm, (await get_header_xh(bbdm)), name, width)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '保存成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '保存失败!'}


async def save_bbgl_filter(bbdm, filter_name, filter_code, filter_type, item, p_item_value, p_notnull, p_is_range,
                           p_rq_range, p_is_like):
    try:
        st = "insert into t_bbgl_filter(bbdm,xh,filter_name,filter_code,filter_type,is_item,item_value,is_null,is_range,rq_range,is_like) \
                 values('{}',{},'{}','{}','{}','{}','{}','{}','{}','{}','{}')". \
            format(bbdm, (await get_filter_xh(bbdm)), filter_name, filter_code, filter_type, item, p_item_value,
                   p_notnull, p_is_range, p_rq_range, p_is_like)
        print(st)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '保存成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '保存失败!'}


async def save_bbgl_preprocess(bbdm, statement, description):
    try:
        st = "insert into t_bbgl_preproccess(bbdm,xh,statement,description) \
                 values('{}',{},'{}','{}')".format(bbdm, (await get_preprocess_xh(bbdm)), fmt_sql(statement),
                                                   description)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '保存成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '保存失败!'}


async def save_bbgl_statement(bbdm, statement):
    try:
        st = """update t_bbgl_config set statement='{}' where bbdm='{}'""".format(fmt_sql(statement), bbdm)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '保存成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '保存失败!'}


async def query_bbgl_header(p_bbdm):
    st = "select xh,header_name,header_width from t_bbgl_header where bbdm='{}' order by bbdm,xh".format(p_bbdm)
    return await async_processer.query_list(st)


async def query_bbgl_filter(p_bbdm):
    st = """select a.xh,
                   a.filter_name,
                   a.filter_code,
                   a.filter_type,
                   b.dmmc,
                   a.is_item,
                   a.item_value,
                   (select mc from t_bbgl_dmlx c where c.dm=a.item_value) as  item_name,
                   a.is_null,a.is_range,a.rq_range,a.is_like 
            from t_bbgl_filter a,t_dmmx b 
             where a.filter_type=b.dmm and b.dm='42'              
              and a.bbdm='{}' order by a.bbdm,a.xh""".format(p_bbdm)
    return await async_processer.query_list(st)


async def query_bbgl_preprocess(p_bbdm):
    st = "select xh,substr(statement,1,50) as statement,description from t_bbgl_preproccess where bbdm='{}' order by bbdm,xh".format(
        p_bbdm)
    return await async_processer.query_list(st)


async def query_bbgl_statement(p_bbdm):
    st = "select statement from t_bbgl_config where bbdm='{}'".format(p_bbdm)
    return await async_processer.query_dict_one(st)


async def check_bbdm_exists(p_bbdm):
    st = "select count(0) as rec from t_bbgl_config where bbdm='{}'".format(p_bbdm)
    print('rs=', (await async_processer.query_one(st))[0])
    return (await async_processer.query_one(st))[0]


async def get_bbgl_bbdm(p_userid):
    st = """select c.bbdm,c.bbmc from t_bbgl_config c
            WHERE EXISTS(
              SELECT dmm FROM t_dict_group_user tu
                WHERE tu.dm=52 AND tu.user_id={}
                  -- AND INSTR(tu.dmm,c.bbdm)>0
                  AND FIND_IN_SET(c.bbdm,tu.dmm) > 0
            ) order by id""".format(p_userid)
    return await async_processer.query_list(st)


async def query_bbgl_data(bbdm, param):
    try:
        # 1.通过bbdm获取报表定义相关数据
        cfg = await get_config(bbdm)
        preprocess = await get_preprocess(bbdm)
        headers = await get_headers(bbdm)

        ds = await get_ds_by_dsid(cfg['dsid'])

        print('param=', param)

        # 2.预处理
        if preprocess != []:
            # 3.获取预处理脚本，替换变量为实参
            for s in preprocess:
                s['replace_statement'] = s['statement']
                if len(param) > 0:
                    for key, value in param.items():
                        s['replace_statement'] = s['replace_statement'].replace('$$' + key + '$$', value)

            # 4.执行预处理代码
            start_time = datetime.datetime.now()
            # for s in preprocess:
            #     print('---->',s['replace_statement'])
            # await async_processer.exec_sql_by_ds(ds,s['replace_statement'])

            # 2023.5.6 optimize pre sql one session execute
            batch_pre_statement = '\n'.join(
                [s['replace_statement'] if s['replace_statement'][-1] == ';' else s['replace_statement'] + ';' for s in
                 preprocess])
            print('batch_pre_statement=', batch_pre_statement)
            await async_processer.exec_sql_by_ds_multi(ds, batch_pre_statement)
            preProcessTime = get_seconds(start_time)
        else:
            preProcessTime = 0

        # 5. 处理查询定义中占位符
        cfg['replace_statement'] = cfg['statement']
        if len(param) > 0:
            for key, value in param.items():
                cfg['replace_statement'] = cfg['replace_statement'].replace('$$' + key + '$$', value)

        # 执行查询
        print('replace_statement=', cfg['replace_statement'])
        result = await exe_query_exp(cfg['dsid'], cfg['replace_statement'], cfg['db'])

        # 替换表头
        if headers != []:
            xh = 0
            for i in result['column']:
                i['title'] = headers[xh]['header_name']
                i['sWidth'] = '{}px'.format(headers[xh]['header_width'])
                xh = xh + 1

        file_time = current_rq()
        if len(param) > 0:
            if param.get('bbrq_begin') and param.get('bbrq_end'):
                file_time = '{}-{}'.format(param.get('bbrq_begin'), param.get('bbrq_end'))
            elif param.get('bbrq'):
                file_time = param.get('bbrq')
            else:
                pass

        result = {"data": result['data'], "column": result['column'], "status": result['status'], "msg": result['msg'],
                  "preTime": str(preProcessTime), "file_time": file_time}
        return result

    except:
        error = traceback.format_exc()[traceback.format_exc().find('pymysql.err.ProgrammingError'):]
        result = {"data": '', "column": '', "status": '1', "msg": error}
        return result


async def update_bbgl_header(p_bbdm, p_xh, p_name, p_width):
    try:
        st = "update t_bbgl_header " \
             " set header_name='{}',header_width='{}' " \
             "  where bbdm='{}' and xh={}".format(p_name, p_width, p_bbdm, p_xh)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '更新成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '更新失败!'}


async def delete_bbgl_header(p_bbdm, p_xh):
    try:
        st = "delete from  t_bbgl_header where  bbdm='{}' and xh={}".format(p_bbdm, p_xh)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '删除成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '删除失败!'}


async def update_bbgl_filter(p_bbdm, p_xh, p_name, p_code, p_type, p_item, p_item_value, p_notnull, p_is_range,
                             p_rq_range, p_is_like):
    try:
        st = "update t_bbgl_filter " \
             " set filter_name='{}',filter_code='{}',filter_type='{}',is_item='{}'," \
             "item_value='{}',is_null='{}',is_range='{}',rq_range='{}',is_like='{}' " \
             "  where bbdm='{}' and xh={}".format(p_name, p_code, p_type, p_item, p_item_value,
                                                  p_notnull, p_is_range, p_rq_range, p_is_like, p_bbdm, p_xh)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '更新成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '更新失败!'}


async def delete_bbgl_filter(p_bbdm, p_xh):
    try:
        st = "delete from  t_bbgl_filter where  bbdm='{}' and xh={}".format(p_bbdm, p_xh)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '删除成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '删除失败!'}


async def query_bbgl_preprocess_detail(p_bbdm, p_xh):
    st = "select * from t_bbgl_preproccess where bbdm='{}' and xh={}".format(p_bbdm, p_xh)
    print(st)
    return await async_processer.query_dict_one(st)


async def update_bbgl_preprocess(p_bbdm, p_xh, p_statement, p_description):
    try:
        st = "update t_bbgl_preproccess " \
             " set statement='{}',`description`='{}' " \
             "  where bbdm='{}' and xh={}".format(fmt_sql(p_statement), p_description, p_bbdm, p_xh)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '更新成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '更新失败!'}


async def delete_bbgl_preprocess(p_bbdm, p_xh):
    try:
        st = "delete from  t_bbgl_preproccess where  bbdm='{}' and xh={}".format(p_bbdm, p_xh)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '删除成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '删除失败!'}


async def update_bbgl_statement(p_bbdm, p_statement):
    try:
        st = "update t_bbgl_config  set statement='{}' where bbdm='{}'".format(fmt_sql(p_statement), p_bbdm)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '更新成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '更新失败!'}


async def query_bbgl_config(p_bbdm,p_bbmc,p_userid):
    vv = ''
    if p_bbdm != '':
        vv = " and a.bbdm='{}'".format(p_bbdm)
    if p_bbmc != '':
        vv = " and instr(a.bbmc,'{}')>0".format(p_bbmc.strip())

    st = """select 
                 bbdm,bbmc,b.id,b.db_desc,a.db,u.name,
                 date_format(a.create_date,'%Y-%m-%d')    create_date,
                 date_format(a.last_update_date,'%Y-%m-%d') last_update_date 
           from t_bbgl_config a,t_db_source b,t_user u
           where a.dsid=b.id and a.creator=u.id {} 
           and EXISTS(
              SELECT dmm FROM t_dict_group_user tu
                WHERE tu.dm=52 AND tu.user_id={}
                 AND FIND_IN_SET(a.bbdm,tu.dmm) > 0
            ) order by id
           """.format(vv,p_userid)
    return await async_processer.query_list(st)


async def query_bbgl_export(p_bbdm,p_userid):
    vv = ''
    if p_bbdm != '':
        vv = " and a.bbdm='{}'".format(p_bbdm)

    st = """SELECT 
                 t.id,
                 a.bbdm,
                 a.bbmc,
                 m.dmmc AS STATUS,
                 t.process,
                 u.name,
                 DATE_FORMAT(t.create_date,'%Y-%m-%d %H:%i:%s')    create_date
            FROM t_bbgl_config a,t_bbgl_export t,t_user u,t_dmmx m
            WHERE a.bbdm=t.bbdm AND t.creator=u.id AND m.dm='43' 
              AND m.dmm=t.status
              and a.bbdm='{}'
              and u.id={}""".format(p_bbdm,p_userid)
    print('st=', st)
    return await async_processer.query_list(st)


async def delete_bbgl(p_bbdm):
    try:
        st = "delete from  t_bbgl_header where  bbdm='{}'".format(p_bbdm)
        await async_processer.exec_sql(st)

        st = "delete from  t_bbgl_preproccess where  bbdm='{}'".format(p_bbdm)
        await async_processer.exec_sql(st)

        st = "delete from  t_bbgl_filter where  bbdm='{}'".format(p_bbdm)
        await async_processer.exec_sql(st)

        st = "delete from  t_bbgl_config where  bbdm='{}'".format(p_bbdm)
        await async_processer.exec_sql(st)

        return {'code': 0, 'message': '删除成功!'}

    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '删除失败!'}


async def export_insert(p_bbdm, p_param, p_userid):
    st = """insert into t_bbgl_export(bbdm,filter,status,process,creator,create_date)
                values('{}','{}','1','0%','{}',now())""".format(p_bbdm, json.dumps(p_param), p_userid)
    id = await async_processer.exec_ins_sql(st)
    return id


async def update_export(p_id, p_status, p_process, p_file='', p_real_file='', p_size='', p_error=''):
    st = "update t_bbgl_export a " \
         " set a.status='{}',a.process='{}',a.file='{}',a.real_file='{}',a.size='{}',a.error='{}' where id={}" \
        .format(p_status, p_process, p_file, p_real_file, p_size, p_error, p_id)
    await async_processer.exec_sql(st)


def set_header_styles(p_fontsize, p_color):
    header_borders = xlwt.Borders()
    header_styles = xlwt.XFStyle()
    # add table header style
    header_borders.left = xlwt.Borders.THIN
    header_borders.right = xlwt.Borders.THIN
    header_borders.top = xlwt.Borders.THIN
    header_borders.bottom = xlwt.Borders.THIN
    header_styles.borders = header_borders
    header_pattern = xlwt.Pattern()
    header_pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    header_pattern.pattern_fore_colour = p_color
    # add font
    font = xlwt.Font()
    font.name = u'微软雅黑'
    font.bold = True
    font.size = p_fontsize
    header_styles.font = font
    # add alignment
    header_alignment = xlwt.Alignment()
    header_alignment.horz = xlwt.Alignment.HORZ_CENTER
    header_alignment.vert = xlwt.Alignment.VERT_CENTER
    header_styles.alignment = header_alignment
    header_styles.borders = header_borders
    header_styles.pattern = header_pattern
    return header_styles


def set_row_styles(p_fontsize, p_color):
    cell_borders = xlwt.Borders()
    cell_styles = xlwt.XFStyle()

    # add font
    font = xlwt.Font()
    font.name = u'微软雅黑'
    font.bold = True
    font.size = p_fontsize
    cell_styles.font = font

    # add col style
    cell_borders.left = xlwt.Borders.THIN
    cell_borders.right = xlwt.Borders.THIN
    cell_borders.top = xlwt.Borders.THIN
    cell_borders.bottom = xlwt.Borders.THIN

    row_pattern = xlwt.Pattern()
    row_pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    row_pattern.pattern_fore_colour = p_color

    # add alignment
    cell_alignment = xlwt.Alignment()
    cell_alignment.horz = xlwt.Alignment.HORZ_LEFT
    cell_alignment.vert = xlwt.Alignment.VERT_CENTER

    cell_styles.alignment = cell_alignment
    cell_styles.borders = cell_borders
    cell_styles.pattern = row_pattern
    cell_styles.font = font
    return cell_styles


async def exp_data(static_path, p_bbdm, p_data, p_id):
    try:
        row_data = 0
        workbook = xlwt.Workbook(encoding='utf8')
        worksheet = workbook.add_sheet(p_bbdm)
        header_styles = set_header_styles(45, 1)
        file_path = static_path + '/downloads/bbtj'
        os.system('cd {0}'.format(file_path))
        file_name = static_path + '/downloads/bbtj/exp_bbtj_{}_{}.xls'.format(p_bbdm, current_rq())
        file_name_s = 'exp_bbtj_{}_{}.xls'.format(p_bbdm, current_rq())

        # write header
        for k in range(len(p_data['column'])):
            worksheet.write(row_data, k, p_data['column'][k]['title'], header_styles)
        await update_export(p_id, '3', '25%')

        # write body
        n_batch_size = 100
        n_total_rows = len(p_data['data'])
        print('n_total_rows=', n_total_rows)
        row_data = row_data + 1
        for i in p_data['data']:
            for j in range(len(i)):
                if i[j] is None:
                    worksheet.write(row_data, j, '')
                else:
                    worksheet.write(row_data, j, str(i[j]))
            row_data = row_data + 1
            if row_data % n_batch_size == 0:
                await update_export(p_id, '3', str(round(row_data / 75, 2) * 100) + '%')

        await update_export(p_id, '3', '98%')

        workbook.save(file_name)
        print("{0} export complete!".format(file_name))

        zip_file = static_path + '/downloads/bbtj/exp_bbtj_{}_{}.zip'.format(p_bbdm, current_rq())
        rzip_file = '/static/downloads/bbtj/exp_bbtj_{}_{}.zip'.format(p_bbdm, current_rq())

        if os.path.exists(zip_file):
            os.system('rm -f {0}'.format(zip_file))

        z = zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED, allowZip64=True)
        z.write(file_name, arcname=file_name_s)
        z.close()

        file_size = os.path.getsize(file_name)
        os.system('rm -f {0}'.format(file_name))

        # update path,file,size
        await update_export(p_id, '3', '100%', rzip_file, zip_file, file_size)
        return rzip_file
    except:
        await update_export(p_id, '34', '0%', '', '', '', traceback.print_exc())
        return ''


async def exp_data_xlsx(static_path, p_bbdm, p_data, p_id):
    try:
        row_data = 1
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(index=0, title=p_bbdm)
        file_path = static_path + '/downloads/bbtj'
        os.system('cd {0}'.format(file_path))
        file_name = static_path + '/downloads/bbtj/exp_bbtj_{}_{}.xlsx'.format(p_bbdm, p_data['file_time'])
        file_name_s = 'exp_bbtj_{}_{}.xls'.format(p_bbdm, p_data['file_time'])

        # write header
        for k in range(len(p_data['column'])):
            ws.cell(column=k + 1, row=row_data, value=p_data['column'][k]['title'])
        await update_export(p_id, '3', '25%')

        # write body
        n_batch_size = 500
        n_total_rows = len(p_data['data'])
        row_data = row_data + 1
        for i in p_data['data']:
            for j in range(len(i)):
                if i[j] is None:
                    ws.cell(row=row_data, column=j + 1, value='')
                else:
                    ws.cell(row=row_data, column=j + 1, value=str(i[j]))
            row_data = row_data + 1
            if row_data % n_batch_size == 0:
                await update_export(p_id, '3', str(round(row_data / 75, 2) * 100) + '%')

        await update_export(p_id, '3', '98%')

        wb.save(file_name)
        print("{0} export complete!".format(file_name))

        zip_file = static_path + '/downloads/bbtj/exp_bbtj_{}_{}.zip'.format(p_bbdm, p_data['file_time'])
        rzip_file = '/static/downloads/bbtj/exp_bbtj_{}_{}.zip'.format(p_bbdm, p_data['file_time'])

        if os.path.exists(zip_file):
            os.system('rm -f {0}'.format(zip_file))

        z = zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED, allowZip64=True)
        z.write(file_name, arcname=file_name_s)
        z.close()

        file_size = os.path.getsize(file_name)
        os.system('rm -f {0}'.format(file_name))

        # update path,file,size
        await update_export(p_id, '3', '100%', rzip_file, zip_file, file_size)
        return rzip_file
    except:
        await update_export(p_id, '34', '0%', '', '', '', traceback.print_exc())


async def export_bbgl_data(bbdm, param, userid, path):
    try:
        id = await export_insert(bbdm, param, userid)
        res = await query_bbgl_data(bbdm, param)
        if res['status'] == '1':
            return {"code": -1, "message": res['msg']}
        await update_export(id, '2', '20%')
        zip_file = await exp_data_xlsx(path, bbdm, res, id)
        return {"code": 0, "message": zip_file}
    except:
        return {"code": -1, "message": traceback.print_exc()}


async def get_download(p_id):
    st = "select *from t_bbgl_export where id={}".format(p_id)
    return await async_processer.query_dict_one(st)


async def get_download_files(p_id,path):
    st = "select file_name,file_data as file from t_bbgl_files where id={}".format(p_id)
    file_name, file_data = await async_processer.query_one(st)
    file_path = f'/static/downloads/bbtj/{file_name}'
    file_path_abs = f'{path}/downloads/bbtj/{file_name}'
    print(f'{file_path},{file_path_abs}')
    with open(file_path_abs, "wb") as file:
        file.write(file_data)
    await async_processer.exec_sql(f"update t_bbgl_files set file_path='{file_path}' where id={p_id}")
    st = "select file_path as file from t_bbgl_files where id={}".format(p_id)
    return await async_processer.query_dict_one(st)


async def get_download_files_new(p_id):
    st = "select file_name,file_data as file from t_bbgl_files where id={}".format(p_id)
    return await async_processer.query_one(st)

async def del_export(p_id):
    try:
        res = await get_download(p_id)
        os.system('rm -f {0}'.format(res.get('real_file')))
        st = "delete from t_bbgl_export where id={}".format(p_id)
        await async_processer.exec_sql(st)
        return {'code': 0, 'message': '删除成功!'}
    except Exception as e:
        traceback.print_exc()
        return {'code': -1, 'message': '删除失败!'}


'''
报表平台-任务管理 
'''


def check_task(p_task):
    result = {}
    result['code'] = '0'
    result['message'] = '验证通过'
    return result


async def query_bbgl_task(p_task_tag, p_userid, p_username):
    if p_username == 'admin':
        v_where = ' '
    else:
        v_where = " and EXISTS(SELECT 1 FROM `t_dict_group_user` b WHERE b.user_id={} AND b.dm='51' AND INSTR(b.dmm,m.type)>0)".format(
            p_userid)

    if p_task_tag != '':
        v_where = " and ( a.task_tag like '%{0}%' or a.comments like '%{1}%' or b.server_ip like '%{2}%' or m.name like '%{3}%')".format(
            p_task_tag, p_task_tag, p_task_tag, p_task_tag)
    sql = """SELECT  
                 task_tag,
                 comments,
                 CONCAT(b.server_ip,':',b.server_port) AS sync_server,             
                 run_time,
                 api_server,
                 CASE a.STATUS WHEN '1' THEN '启用' WHEN '0' THEN '禁用' END  AS  flag
            FROM t_bbgl_task a,t_server b
            where a.server_id=b.id 
            {0}""".format(v_where)
    print(sql)
    return await async_processer.query_list(sql)


async def save_bbgl_task(p_task):
    val = check_task(p_task)
    if val['code'] == '-1':
        return val
    try:
        sql = """insert into t_bbgl_task (task_tag,comments,server_id,bbid,tjrq_begin,tjrq_end,tjrq_begin_type,
                    tjrq_end_type,run_time,script_path,script_file,python3_home,api_server,status,receiver,cc)
                      values('{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}','{11}','{12}','{13}','{14}','{15}')
             """.format(p_task['add_bbgl_task_tag'],
                        p_task['add_bbgl_task_desc'],
                        p_task['add_bbgl_server'],
                        p_task['add_bbgl_id'],
                        p_task['add_bbgl_tjrq_begin_value'],
                        p_task['add_bbgl_tjrq_end_value'],
                        p_task['add_bbgl_tjrq_begin_type'],
                        p_task['add_bbgl_tjrq_end_type'],
                        p_task['add_bbgl_task_run_time'],
                        format_sql(p_task['add_bbgl_task_script_base']),
                        format_sql(p_task['add_bbgl_task_script_name']),
                        format_sql(p_task['add_bbgl_task_python3_home']),
                        p_task['add_bbgl_task_api_server'],
                        p_task['add_bbgl_task_status'],
                        p_task['add_bbgl_receiver'],
                        p_task['add_bbgl_cc'])
        await async_processer.exec_sql(sql)
        return {'code': '0', 'message': '保存成功!'}
    except:
        traceback.print_exc()
        return {'code': '-1', 'message': '保存失败!'}


async def upd_bbgl_task(p_task):
    val = check_task(p_task)
    if val['code'] == '-1':
        return val
    try:
        sql = """update t_bbgl_task
                  set  task_tag='{}',
                       comments='{}',
                       run_time='{}',
                       script_path='{}',
                       script_file='{}',
                       python3_home='{}',
                       api_server='{}',
                       status='{}',
                       bbid ='{}',
                       tjrq_begin='{}',
                       tjrq_end='{}',
                       tjrq_begin_type='{}',                     
                       tjrq_end_type='{}',
                       receiver='{}',
                       cc='{}',
                       server_id='{}'
                 where task_tag='{}'
             """.format(p_task['upd_bbgl_task_tag'],
                        p_task['upd_bbgl_task_desc'],
                        p_task['upd_bbgl_task_run_time'],
                        format_sql(p_task['upd_bbgl_task_script_base']),
                        format_sql(p_task['upd_bbgl_task_script_name']),
                        format_sql(p_task['upd_bbgl_task_python3_home']),
                        p_task['upd_bbgl_task_api_server'],
                        p_task['upd_bbgl_task_status'],
                        p_task['upd_bbgl_id'],
                        p_task['upd_bbgl_tjrq_begin_value'],
                        p_task['upd_bbgl_tjrq_end_value'],
                        p_task['upd_bbgl_tjrq_begin_type'],
                        p_task['upd_bbgl_tjrq_end_type'],
                        p_task['upd_bbgl_receiver'],
                        p_task['upd_bbgl_cc'],
                        p_task['upd_bbgl_server'],
                        p_task['upd_bbgl_task_tag_old'])
        print('sql=', sql)
        await async_processer.exec_sql(sql)
        return {'code': '0', 'message': '保存成功!'}
    except:
        traceback.print_exc()
        return {'code': '-1', 'message': '保存失败!'}


async def del_bbgl_task(p_task_tag):
    try:
        sql = "delete from t_bbgl_task  where task_tag='{0}'".format(p_task_tag)
        await async_processer.exec_sql(sql)
        return {'code': '0', 'message': '删除成功!'}
    except:
        traceback.print_exc()
        return {'code': '-1', 'message': '删除失败!'}


def push_bbgl_task(p_tag, p_api):
    url = 'http://{}/push_script_remote_bbgl'.format(p_api)
    res = requests.post(url, data={'tag': p_tag})
    jres = res.json()
    if jres['code'] == 200:
        v = ''
        for c in jres['msg']:
            if c.count(p_tag) > 0:
                v = v + "<span class='warning'>" + c + "</span>"
            else:
                v = v + c
            v = v + '<br>'
        jres['msg'] = v
        return jres
    else:
        return jres


def run_bbgl_task(p_tag, p_api):
    url = 'http://{}/run_script_remote_bbgl'.format(p_api)
    res = requests.post(url, data={'tag': p_tag})
    jres = res.json()
    return jres


def stop_bbgl_task(p_tag, p_api):
    url = 'http://{}/stop_script_remote_bbgl'.format(p_api)
    res = requests.post(url, data={'tag': p_tag})
    jres = res.json()
    return jres


async def get_bbgl_task_by_tag(p_tag):
    sql = """SELECT  * FROM t_bbgl_task where task_tag='{0}'""".format(p_tag)
    return (await async_processer.query_dict_one(sql))


async def get_bbgl_id():
    sql = "select id,bbmc from t_bbgl_config order by 1"
    return await async_processer.query_list(sql)


async def get_bbgl_tjrq_type():
    sql = "select dmm,dmmc from t_bbgl_dmmx where dm='002' order by dm"
    return await async_processer.query_list(sql)


async def get_bbgl_tjrq_value(tjlx):
    sql = "select dmm,dmmc,dmmc2 from t_bbgl_dmmx where dm='{}' order by dm+0".format(tjlx)
    return await async_processer.query_list(sql)


async def get_imp_data_type():
    sql = "select dmm,dmmc from t_bbgl_dmmx where dm='009' order by dm"
    return await async_processer.query_list(sql)
